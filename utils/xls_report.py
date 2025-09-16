import csv
import os
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def generate_csv_report(organized_files: dict, destination_directory: str, report_type: str):
    """
    Generate a CSV report for audio or other files.
    For audio: columns = title, artist, album, file path
    For other: columns = file name, directory, last modified date
    """
    if report_type == 'audio':
        report_path = os.path.join(destination_directory, 'audio', 'audio_report.csv')
        rows = []
        for src, dest in organized_files.get('audio', {}).items():
            meta = organized_files.get('audio_metadata', {}).get(src, {})
            rows.append([
                meta.get('title', 'UnknownTitle'),
                meta.get('artist', 'UnknownArtist'),
                meta.get('album', 'UnknownAlbum'),
                dest
            ])
        with open(report_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Title', 'Artist', 'Album', 'File Path'])
            writer.writerows(rows)
        return report_path
    elif report_type == 'other':
        report_path = os.path.join(destination_directory, 'other', 'other_report.csv')
        rows = []
        for src, dest in organized_files.get('other', {}).items():
            fname = os.path.basename(src)
            directory = os.path.dirname(src)
            mtime = os.path.getmtime(src)
            rows.append([fname, directory, mtime])
        with open(report_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['File Name', 'Directory', 'Last Modified Date'])
            writer.writerows(rows)
        return report_path
    else:
        raise ValueError('Unknown report type')

def generate_excel_report(organized_files: dict, destination_directory: str):
    """
    Generate an Excel (.xlsx) report for 'other' files, showing file name, destination path (without file name), and last modified date (Excel date format).
    Columns are wide enough to avoid text wrapping.
    """
    report_path = os.path.join(destination_directory, 'other', 'other_report.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Other Files'
    ws.append(['File Name', 'Destination Path', 'Last Modified Date'])
    for src, dest in organized_files.get('other', {}).items():
        fname = os.path.basename(src)
        dest_path = os.path.dirname(dest)
        mtime = os.path.getmtime(dest) if os.path.exists(dest) else os.path.getmtime(src)
        excel_date = datetime.datetime.fromtimestamp(mtime)
        ws.append([fname, dest_path, excel_date])
    # Format columns
    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(2)].width = 60
    ws.column_dimensions[get_column_letter(3)].width = 20
    for cell in ws[get_column_letter(3)][1:]:
        cell.number_format = 'mm/dd/yyyy'
    wb.save(report_path)
    return report_path

def generate_master_report(file_metadata: list, categorized_files: dict, copied_files: dict, destination_directory: str):
    """
    Generate a master Excel report with file name, source path, destination path, and result (Copied, Ignored, Renamed).
    If ignored, destination path is N/A.
    Columns are wide enough to avoid text wrapping.
    """
    report_path = os.path.join(destination_directory, 'master_report.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Master Report'
    ws.append(['File Name', 'Source Path', 'Destination Path', 'Result'])
    all_dest_paths = set()
    for category, files in copied_files.items():
        all_dest_paths.update(files.values())
    for meta in file_metadata:
        fname = os.path.basename(meta['path'])
        src_path = meta['path']
        dest_path = 'N/A'
        result = 'Ignored'
        for category, files in categorized_files.items():
            if src_path in files:
                candidate_dest = files[src_path]
                if candidate_dest in all_dest_paths:
                    dest_path = candidate_dest
                    if fname != os.path.basename(dest_path):
                        result = 'Renamed'
                    else:
                        result = 'Copied'
                else:
                    dest_path = 'N/A'
                    result = 'Ignored'
                break
        ws.append([fname, src_path, dest_path, result])
    # Format columns
    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(2)].width = 60
    ws.column_dimensions[get_column_letter(3)].width = 60
    ws.column_dimensions[get_column_letter(4)].width = 15
    wb.save(report_path)
    return report_path
